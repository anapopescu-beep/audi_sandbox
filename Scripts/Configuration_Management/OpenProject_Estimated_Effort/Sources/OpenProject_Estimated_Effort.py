import openpyxl
from openpyxl import load_workbook, Workbook
import argparse
import math
from tqdm import tqdm


def sum_for_one_row(current_row, estimated_hours, task, path_excel):

    # Opening the input Excel file
    try:
        path = path_excel
        input_wb = load_workbook(path, read_only=True, data_only=True)
    except:
        print("Could not load Excel input file")

    sheet = input_wb["Detailed estimation"]

    # Based on the current row, parse through the Tasks
    task_current_row = current_row - 1
    for task_row in sheet.iter_rows(min_row=current_row, min_col=4, max_col=4, max_row=current_row + 42):
        task_current_row += 1
        for task_cell in task_row:

            # For the task that we need, add all estimated depending on the state of the task
            if task_cell.value == task:
                if sheet.cell(current_row, 23).value == "Concept" and sheet.cell(task_current_row, 9).value is not None:
                    estimated_hours = estimated_hours + sheet.cell(task_current_row, 9).value
                if sheet.cell(current_row, 23).value == "R1" and sheet.cell(task_current_row, 10).value is not None:
                    estimated_hours = estimated_hours + sheet.cell(task_current_row, 10).value
                if sheet.cell(current_row, 23).value == "R1.1" and sheet.cell(task_current_row, 11).value is not None:
                    estimated_hours = estimated_hours + sheet.cell(task_current_row, 11).value
                if sheet.cell(current_row, 23).value == "R1.2" and sheet.cell(task_current_row, 12).value is not None:
                    estimated_hours = estimated_hours + sheet.cell(task_current_row, 12).value
                if sheet.cell(current_row, 23).value == "R2" and sheet.cell(task_current_row, 13).value is not None:
                    estimated_hours = estimated_hours + sheet.cell(task_current_row, 13).value
                if sheet.cell(current_row, 23).value == "R2.1" and sheet.cell(task_current_row, 14).value is not None:
                    estimated_hours = estimated_hours + sheet.cell(task_current_row, 14).value
                if sheet.cell(current_row, 23).value == "R2.2" and sheet.cell(task_current_row, 15).value is not None:
                    estimated_hours = estimated_hours + sheet.cell(task_current_row, 15).value
                if sheet.cell(current_row, 23).value == "R3" and sheet.cell(task_current_row, 16).value is not None:
                    estimated_hours = estimated_hours + sheet.cell(task_current_row, 16).value
                if sheet.cell(current_row, 23).value == "R3.1" and sheet.cell(task_current_row, 17).value is not None:
                    estimated_hours = estimated_hours + sheet.cell(task_current_row, 17).value
                if sheet.cell(current_row, 23).value == "R3.2" and sheet.cell(task_current_row, 18).value is not None:
                    estimated_hours = estimated_hours + sheet.cell(task_current_row, 18).value
    return estimated_hours

def process_estimated_hours(input_path, output_path):

    # Open the Excel input file
    try:
        path = input_path
        input_wb = load_workbook(path, read_only=True, data_only=True)
    except:
        print("Could not load Excel input file")

    sheet = input_wb["Detailed estimation"]

    # Find the position of the OpenProject Task ID column
    current_row = 0
    for row in sheet.iter_rows(min_row=1, min_col=1):
        current_col = 0
        current_row += 1
        for cell in row:
            current_col += 1
            if cell.value is not None and cell.value == "OpenProject Task ID3":
                project_id_row = current_row
                project_id_col = current_col

    estimated_hours_dict = {}

    print("Reading data from Excel")
    print("==========================================")

    # Parse through the OpenProject Task ID column
    current_row = project_id_row
    for row in sheet.iter_rows(min_row=project_id_row+1, min_col=project_id_col, max_col=project_id_col):
        current_row += 1
        for cell in row:
            issue_id = cell.value
            # if issue_id is not None and type(issue_id) == int:
            #     estimated_hours_list = []
            #     analysis_hours = 0
            #     requirements_hours = 0
            #     architecture_hours = 0
            #     design_hours = 0
            #     code_implementation_hours = 0
            #     static_hours = 0
            #     unit_test_hours = 0
            #     integration_test_hours = 0
            #     review_hours = 0
            #
            #     estimated_hours_list.append(sum_for_one_row(current_row, requirements_hours, "SW Requirements Review", input_path))
            #     estimated_hours_list.append(sum_for_one_row(current_row, architecture_hours, "SW Architecture Specification", input_path))
            #     estimated_hours_list.append(sum_for_one_row(current_row, design_hours, "Design Specification", input_path))
            #     estimated_hours_list.append(sum_for_one_row(current_row, code_implementation_hours, "Code Implementation", input_path))
            #     # estimated_hours_list.append(sum_for_one_row(current_row, static_hours, "QAC/Code Metrics Verification", input_path))
            #     estimated_hours_list.append(sum_for_one_row(current_row, unit_test_hours, "Unit Test Specification", input_path) +
            #                                 sum_for_one_row(current_row, unit_test_hours, "Unit Test Execution", input_path) +
            #                                 sum_for_one_row(current_row, static_hours, "QAC/Code Metrics Verification", input_path))
            #     estimated_hours_list.append(sum_for_one_row(current_row, integration_test_hours, "SW Integration Test Specification", input_path) +
            #                                 sum_for_one_row(current_row, integration_test_hours, "SW Integration Test Execution", input_path))
            #     estimated_hours_list.append(sum_for_one_row(current_row, review_hours, "SW Architecture Specification Review", input_path) +
            #                                 sum_for_one_row(current_row, review_hours, "Design Specification Review", input_path) +
            #                                 sum_for_one_row(current_row, review_hours, "Code Review", input_path) +
            #                                 sum_for_one_row(current_row, review_hours, "Unit Test Specification Review", input_path) +
            #                                 sum_for_one_row(current_row, review_hours, "SW Integration Test Specification Review", input_path))
            #     estimated_hours_list.append(sum_for_one_row(current_row, analysis_hours, "Customer Req Analysis", input_path))
            #
            #     sum = 0
            #     for elem in estimated_hours_list:
            #         sum = sum + elem
            #
            #     story_points = sum / 4
            #     final_story_point = math.ceil(story_points)
            #     estimated_hours_list.append(final_story_point)
            #
            #     estimated_hours_dict.update(
            #         {cell.value: estimated_hours_list
            #          }
            #     )
            #
            # elif issue_id is not None and type(issue_id) == str and sheet.cell(current_row, 3).value != "SW Bug-Fix":
            #
            #     issue_id = issue_id.split(";")
            #     for elem in range(1, len(issue_id)):
            #         if "%" in issue_id[elem]:
            #             issue_id[elem] = issue_id[elem].replace("%", "")
            #     it = iter(issue_id)
            #     for x in it:
            #         estimated_hours_list = []
            #         analysis_hours = 0
            #         requirements_hours = 0
            #         architecture_hours = 0
            #         design_hours = 0
            #         code_implementation_hours = 0
            #         static_hours = 0
            #         unit_test_hours = 0
            #         integration_test_hours = 0
            #         review_hours = 0
            #         single_id = int(x)
            #         single_percentage = int(next(it))
            #
            #         sum_requirements_hours = sum_for_one_row(current_row, requirements_hours, "SW Requirements Review", input_path) / 100 * single_percentage
            #         sum_requirements_hours = int(sum_requirements_hours)
            #         if single_percentage > 50:
            #             sum_requirements_hours = sum_requirements_hours + 1
            #         estimated_hours_list.append(sum_requirements_hours)
            #         sum_architecture_hours = sum_for_one_row(current_row, architecture_hours, "SW Architecture Specification", input_path) / 100 * single_percentage
            #         sum_architecture_hours = int(sum_architecture_hours)
            #         if single_percentage > 50:
            #             sum_architecture_hours = sum_architecture_hours + 1
            #         estimated_hours_list.append(sum_architecture_hours)
            #         sum_design_hours = sum_for_one_row(current_row, design_hours, "Design Specification", input_path) / 100 * single_percentage
            #         sum_design_hours = int(sum_design_hours)
            #         if single_percentage > 50:
            #             sum_design_hours = sum_design_hours + 1
            #         estimated_hours_list.append(sum_design_hours)
            #         sum_code_implementation_hours = sum_for_one_row(current_row, code_implementation_hours, "Code Implementation", input_path) / 100 * single_percentage
            #         sum_code_implementation_hours = int(sum_code_implementation_hours)
            #         if single_percentage > 50:
            #             sum_code_implementation_hours = sum_code_implementation_hours + 1
            #         estimated_hours_list.append(sum_code_implementation_hours)
            #         sum_unit_test_hours = (
            #             sum_for_one_row(current_row, unit_test_hours, "Unit Test Specification", input_path) +
            #             sum_for_one_row(current_row, unit_test_hours, "Unit Test Execution", input_path) +
            #             sum_for_one_row(current_row, static_hours, "QAC/Code Metrics Verification", input_path)) / 100 * single_percentage
            #         sum_unit_test_hours = int(sum_unit_test_hours)
            #         if single_percentage > 50:
            #             sum_unit_test_hours = sum_unit_test_hours + 1
            #         estimated_hours_list.append(sum_unit_test_hours)
            #         sum_integration_hours = (
            #             sum_for_one_row(current_row, integration_test_hours, "SW Integration Test Specification", input_path) +
            #             sum_for_one_row(current_row, integration_test_hours, "SW Integration Test Execution", input_path)) / 100 * single_percentage
            #         sum_integration_hours = int(sum_integration_hours)
            #         if single_percentage > 50:
            #             sum_integration_hours = sum_integration_hours + 1
            #         estimated_hours_list.append(sum_integration_hours)
            #         sum_review_hours = (
            #             sum_for_one_row(current_row, review_hours, "SW Architecture Specification Review", input_path) +
            #             sum_for_one_row(current_row, review_hours, "Design Specification Review", input_path) +
            #             sum_for_one_row(current_row, review_hours, "Code Review", input_path) +
            #             sum_for_one_row(current_row, review_hours, "Unit Test Specification Review", input_path) +
            #             sum_for_one_row(current_row, review_hours, "SW Integration Test Specification Review", input_path)) / 100 * single_percentage
            #         sum_review_hours = int(sum_review_hours)
            #         if single_percentage > 50:
            #             sum_review_hours = sum_review_hours + 1
            #         estimated_hours_list.append(sum_review_hours)
            #         sum_analysis_hours = sum_for_one_row(current_row, analysis_hours, "Customer Req Analysis", input_path) / 100 * single_percentage
            #         sum_analysis_hours = int(sum_analysis_hours)
            #         if single_percentage > 50:
            #             sum_analysis_hours = sum_analysis_hours + 1
            #         estimated_hours_list.append(sum_analysis_hours)
            #
            #         sum = 0
            #         for elem in estimated_hours_list:
            #             sum = sum + elem
            #
            #         story_points = sum / 4
            #         final_story_point = math.ceil(story_points)
            #         estimated_hours_list.append(final_story_point)
            #
            #         estimated_hours_dict.update(
            #             {single_id: estimated_hours_list
            #              }
            #         )

            if issue_id is not None and type(issue_id) == str and sheet.cell(current_row, 3).value == "SW Bug-Fix":
                issue_id = issue_id.split(";")
                estimated_hours_list = []
                analysis_hours = 0
                requirements_hours = 0
                architecture_hours = 0
                design_hours = 0
                code_implementation_hours = 0
                static_hours = 0
                unit_test_hours = 0
                integration_test_hours = 0
                review_hours = 0

                single_percentage = len(issue_id)

                sum_requirements_hours = sum_for_one_row(current_row, requirements_hours, "SW Requirements Review", input_path) / 100 * single_percentage
                sum_requirements_hours = int(sum_requirements_hours)
                if single_percentage > 50:
                    sum_requirements_hours = sum_requirements_hours + 1
                estimated_hours_list.append(sum_requirements_hours)
                sum_architecture_hours = sum_for_one_row(current_row, architecture_hours, "SW Architecture Specification", input_path) / 100 * single_percentage
                sum_architecture_hours = int(sum_architecture_hours)
                if single_percentage > 50:
                    sum_architecture_hours = sum_architecture_hours + 1
                estimated_hours_list.append(sum_architecture_hours)
                sum_design_hours = sum_for_one_row(current_row, design_hours, "Design Specification", input_path) / 100 * single_percentage
                sum_design_hours = int(sum_design_hours)
                if single_percentage > 50:
                    sum_design_hours = sum_design_hours + 1
                estimated_hours_list.append(sum_design_hours)
                sum_code_implementation_hours = sum_for_one_row(current_row, code_implementation_hours, "Code Implementation", input_path) / 100 * single_percentage
                sum_code_implementation_hours = int(sum_code_implementation_hours)
                if single_percentage > 50:
                    sum_code_implementation_hours = sum_code_implementation_hours + 1
                estimated_hours_list.append(sum_code_implementation_hours)
                sum_unit_test_hours = (
                    sum_for_one_row(current_row, unit_test_hours, "Unit Test Specification", input_path) +
                    sum_for_one_row(current_row, unit_test_hours, "Unit Test Execution", input_path) +
                    sum_for_one_row(current_row, static_hours, "QAC/Code Metrics Verification", input_path)) / 100 * single_percentage
                sum_unit_test_hours = int(sum_unit_test_hours)
                if single_percentage > 50:
                    sum_unit_test_hours = sum_unit_test_hours + 1
                estimated_hours_list.append(sum_unit_test_hours)
                sum_integration_hours = (
                    sum_for_one_row(current_row, integration_test_hours, "SW Integration Test Specification", input_path) +
                    sum_for_one_row(current_row, integration_test_hours, "SW Integration Test Execution", input_path)) / 100 * single_percentage
                sum_integration_hours = int(sum_integration_hours)
                if single_percentage > 50:
                    sum_integration_hours = sum_integration_hours + 1
                estimated_hours_list.append(sum_integration_hours)
                sum_review_hours = (
                    sum_for_one_row(current_row, review_hours, "SW Architecture Specification Review", input_path) +
                    sum_for_one_row(current_row, review_hours, "Design Specification Review", input_path) +
                    sum_for_one_row(current_row, review_hours, "Code Review", input_path) +
                    sum_for_one_row(current_row, review_hours, "Unit Test Specification Review", input_path) +
                    sum_for_one_row(current_row, review_hours, "SW Integration Test Specification Review", input_path)) / 100 * single_percentage
                sum_review_hours = int(sum_review_hours)
                if single_percentage > 50:
                    sum_review_hours = sum_review_hours + 1
                estimated_hours_list.append(sum_review_hours)
                sum_analysis_hours = sum_for_one_row(current_row, analysis_hours, "Customer Req Analysis", input_path) / 100 * single_percentage
                sum_analysis_hours = int(sum_analysis_hours)
                if single_percentage > 50:
                    sum_analysis_hours = sum_analysis_hours + 1
                estimated_hours_list.append(sum_analysis_hours)

                sum = 0
                for elem in estimated_hours_list:
                    sum = sum + elem

                story_points = sum / 4
                final_story_point = math.ceil(story_points)
                estimated_hours_list.append(final_story_point)

                for elem in issue_id:
                    elem_int = int(elem)
                    estimated_hours_dict.update(
                        {elem_int: estimated_hours_list
                         }
                    )


    print("Writing data to Excel template")

    # Open the output Excel template
    try:
        path1 = output_path
        output_wb = openpyxl.load_workbook(path1, keep_vba=True)
    except:
        print("Could not load output excel")

    template_sheet = output_wb['Workpackages']

    max_row = template_sheet.max_row

    # Find matching IDs in the output template and write their hours on their corresponding columns
    for row in range(1, max_row + 1):
        id = template_sheet.cell(row, 3).value
        for elem in estimated_hours_dict.keys():
            if id == elem:
                list = estimated_hours_dict[elem]
                hours_col = 5
                for hours_value in list:
                    template_sheet.cell(row, hours_col).value = hours_value
                    hours_col += 1

    output_wb.save(output_path)


def main():

    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input-path-excel-file', help="Path to the input excel file with esimated hours.", required=True)
    parser.add_argument('-o', '--output-path-excel-file', help="Path to the output excel template file.", required=True)
    args = parser.parse_args()

    path_to_input_excel = args.input_path_excel_file
    path_to_output_file = args.output_path_excel_file

    process_estimated_hours(path_to_input_excel, path_to_output_file)

    print("==========================================")
    print("Script ran successfully")
    print("==========================================")

if __name__ == "__main__":
    main()