import re
from builtins import print
import time

import openpyxl
import xlwt
import pandas as pd
from os import walk
import os
import fnmatch
import sys
from colorama import Fore, Back, Style

currentDir = os.getcwd()
sys.path.append(currentDir)

def parse_integration_tests():
    int_test_file_list = []
    int_test_file_list_statistics = {}
    for root, dirnames, filenames in os.walk("S:\\Tests\\Tests_Integration\\Automatic_Tests\\Test_Implementation\\"):
        for filename in fnmatch.filter(filenames, '*.can'):
            with open(os.path.join(root, filename), "r", encoding='utf-8', errors='ignore') as int_test_file:
                int_test_export_content = int_test_file.readlines()
                lastTestcaseName = ''
                for line in int_test_export_content:
                    if line.strip().startswith('testcase'):
                        lastTestcaseName = line.strip().split(' ')[1].split('(')[0].strip()

                    if '"Requirement under test"' in line:
                        reqList = line.split('"Requirement under test"')[1].split('"')[1].strip().replace(' ', '').split(',')
                        while '' in reqList:
                            reqList.remove('')

                        if lastTestcaseName not in int_test_file_list_statistics.keys():
                            int_test_file_list_statistics[lastTestcaseName] = []
                        int_test_file_list_statistics[lastTestcaseName] += reqList

    # for root, dirnames, filenames in os.walk("S:\\Tests\\Tests_Integration\\Automatic_Tests\\Test_Scenarios\\"):
    #     for filename in fnmatch.filter(filenames, '*.txt'):
    #         int_test_file_list.append(os.path.join(root, filename))
    #
    # for file in int_test_file_list:
    #     int_test_file_list_statistics[os.path.basename(file)] = []
    #     with open(file, "r", encoding='utf-8', errors='ignore') as int_test_file:
    #         int_test_export_content = int_test_file.readlines()
    #
    #     for line in int_test_export_content:
    #         arch_req_match = re.match(r'.*[\s]*(ARCH_SW_[\w, ]*)', line)
    #
    #         if arch_req_match is not None:
    #             temp_arch_str = arch_req_match.group(1).replace(';', ",").replace(' ', '').replace('\t', '')
    #             temp_arch_str = temp_arch_str.split(',')
    #             while '' in temp_arch_str:
    #                 temp_arch_str.remove('')
    #             for temp in temp_arch_str:
    #                 int_test_file_list_statistics[os.path.basename(file)].append(temp)

    return int_test_file_list_statistics

def parse_unit_tests():
    unit_test_file_list_statistics = {}
    unit_test_file_list = []

    for root, dirnames, filenames in os.walk("S:\\Components\\Application\\"):
        for filename in fnmatch.filter(filenames, 'TESSY_*.html'):
            unit_test_file_list.append(os.path.join(root, filename))


    for file in unit_test_file_list:
        unit_test_file_list_statistics[os.path.basename(file)] = []
        with open(file, "r", encoding='utf-8', errors='ignore') as unit_test_file:
            unit_test_export_content = unit_test_file.readlines()

        for line in unit_test_export_content:
            dd_req_match = re.match(r'.*[\s]*(DSG_[\w, ]*)', line)

            if dd_req_match is not None:
                temp_dom_str = dd_req_match.group(1).replace(';', ",").replace(' ', '').replace('\t', '')
                temp_dom_str = temp_dom_str.split(',')
                while '' in temp_dom_str:
                    temp_dom_str.remove('')
                for temp in temp_dom_str:
                    if temp not in unit_test_file_list_statistics[os.path.basename(file)]:
                        unit_test_file_list_statistics[os.path.basename(file)].append(temp)

    return unit_test_file_list_statistics


def parse_MBD_unit_tests():
    unit_test_file_list_statistics = {}
    unit_test_file_list = []

    for root, dirnames, filenames in os.walk("s:\\Components\\Application\\Autoliv\\eCS\\Tests\\Unit_Tests\\"):
        for filename in filenames:
            if filename.endswith('.xlsx') and 'simulation_results_traceability' in filename.lower():
                unit_test_file_list.append(os.path.join(root, filename))

    for file in unit_test_file_list:
        # Open workbook
        wb = openpyxl.load_workbook(file)
        sheets = wb.sheetnames

        # open sheet
        # Variable to check if code traceability sheet is found
        traceability_found_flag = False
        for sheet_name in sheets:
            if 'traceability' in sheet_name.lower():
                ws = wb[sheet_name]
                traceability_found_flag = True
                break
            else:
                traceability_found_flag = False
        if not traceability_found_flag:
            return 'TRACEABILITY SHEET NOT FOUND'
        else:
            # append document data in list
            for i in range(2, ws.max_row):
                testCase = ws.cell(row=i, column=2).value
                DSGreq = ws.cell(row=i, column=1).value
                if testCase not in unit_test_file_list_statistics.keys():
                    unit_test_file_list_statistics[testCase] = []
                unit_test_file_list_statistics[testCase].append(DSGreq)

    return unit_test_file_list_statistics


def parse_source_code():
    c_file_list_statistics = {}
    c_file_list = []

    for root, dirnames, filenames in os.walk('S:/Components/Application/Autoliv/'):
        for filename in fnmatch.filter(filenames, '*.c'):
            if not '_Old.c' in filename and not 'Targetlink' in filename and not '- Copy' in filename and not '\\MBD\\' in root:
                c_file_list.append(os.path.join(root, filename))

    for file in c_file_list:
        with open(file, "r", encoding='utf-8', errors='ignore') as c_export:
            c_export_content = c_export.readlines()
        last_dsg_req = ""

        dsg_req_disable = False
        for line in c_export_content:
            dsg_req_match = re.match(r'.*[\s]*(DSG_[\w, ]*)', line)
            arch_req_match = re.match(r'.*[\s]*(ARCH_SW_[\w, ]*)', line)

            if dsg_req_match is not None:
                if '!DSG_' in line.replace(' ', ''):
                    dsg_req_disable = True
                else:
                    dsg_req_disable = False

                temp_dom_str = dsg_req_match.group(1).replace(';', ",").replace(' ', '').replace('\t', '')
                temp_dom_str = temp_dom_str.split(',')
                while '' in temp_dom_str:
                    temp_dom_str.remove('')
                for temp in temp_dom_str:
                    last_dsg_req = temp
                if last_dsg_req not in c_file_list_statistics.keys() and not dsg_req_disable:
                    c_file_list_statistics[last_dsg_req] = []
            if arch_req_match is not None and not dsg_req_disable:
                temp_dom_str = arch_req_match.group(1).replace(';', ",").replace(' ', '').replace('\t', '')
                temp_dom_str = temp_dom_str.split(',')
                while '' in temp_dom_str:
                    temp_dom_str.remove('')
                for temp in temp_dom_str:
                    if temp not in c_file_list_statistics[last_dsg_req]:
                        c_file_list_statistics[last_dsg_req].append(temp)
    return c_file_list_statistics


def parse_xdm_files():

    xdm_statistics = {}
    xdm_file_list = []

    for root, dirnames, filenames in os.walk('s:\\Components\\Application\\Supplier\\Tresos_Configuration_8.5.1\\Workspace\\Application\\config\\'):
        for filename in fnmatch.filter(filenames, '*.xdm'):
            if not 'bak' in filename:
                xdm_file_list.append(os.path.join(root, filename))

    for root, dirnames, filenames in os.walk('S:\\Components\\Bootloader\\Supplier\\Tresos_Configuration_Bootloader\\Workspace\\demoBoot_VAG_SigCompCRC_can_asr\config\\'):
        for filename in fnmatch.filter(filenames, '*.xdm'):
            if not 'bak' in filename:
                xdm_file_list.append(os.path.join(root, filename))

    for file in xdm_file_list:
        with open(file, "r", encoding='utf-8', errors='ignore') as xdm_export:
            xdm_export_content = xdm_export.readlines()
        filename = os.path.basename(file) + '-' + file.split('Components\\')[1].split('\\')[0]
        xdm_statistics[filename] = []
        for line in xdm_export_content:
            if 'DOM_' in line:
                pass
            match_dom_req = re.match(r'(.*\]:)([ ,]*DOM_eCS[^<>]*)([<>\s\S])', line)

            if match_dom_req is not None:
                temp_dom_str = match_dom_req.group(2).replace(';', ",").replace(' ', '').replace('\t', '')
                temp_dom_str = temp_dom_str.split(',')
                while '' in temp_dom_str:
                    temp_dom_str.remove('')
                for temp in temp_dom_str:
                    xdm_statistics[filename].append(temp)

            match_fusa_req = re.match(r'(.*\]:)([ ,]*FUSA_[^<>]*)([<>\s\S])', line)

            if match_fusa_req is not None:
                temp_fusa_str = match_fusa_req.group(2).replace(';', ",").replace(' ', '').replace('\t', '')
                temp_fusa_str = temp_fusa_str.split(',')
                while '' in temp_fusa_str:
                    temp_fusa_str.remove('')
                for temp in temp_fusa_str:
                    xdm_statistics[filename].append(temp)

    return xdm_statistics


def parse_doors_export():
    doors_req_statistics = {}
    doors_export_file_list = []
    for (dir_path, dir_names, file_names) in walk('./'):
        for file_name in file_names:
            if "eCS_TF_" in file_name:
                doors_export_file_list.append(file_name)
        break

    for file in doors_export_file_list:
        doors_xlsx_file = pd.ExcelFile('./' + file)
        doors_export = pd.read_excel(doors_xlsx_file, 'Sheet1')
        doors_export = doors_export.fillna('')
        doors_export.replace({False: "False", True: 'True'}, inplace=True)

        for row in range(0, len(doors_export['ID'])):
            if type(doors_export['_Type'][row]) is str:
                if 'Functional Requirement' in doors_export['_Type'][row] or 'Non-functional Requirement' in doors_export['_Type'][row]:
                    if doors_export['_Implemented'][row] != "":
                        temp_str = doors_export['_Implemented'][row]
                    else:
                        temp_str = "False"
                    if doors_export['_Target_eCS'][row] != "":
                        temp_str2 = doors_export['_Target_eCS'][row]
                    else:
                        temp_str2 = "Not allocated"
                    doors_req_statistics[doors_export['ID'][row]] = [temp_str2, temp_str, doors_export['_ReqStatus'][row], doors_export['_ImpactedDiscipline'][row], doors_export['_ASIL'][row]]

    return doors_req_statistics


def parse_nvp_files():
    nvpFiles = ["s:\\Components\\Application\\Autoliv\\NVP\\Config\\NVP_Layout\\NvmDefault.xml", "s:\\Components\\Application\\Autoliv\\NVP\\Config\\NVP_Layout\\Calibration.xml"]
    nvp_export_traceability = {}
    for nvpFile in nvpFiles:
        with open(nvpFile, "r", encoding='utf-8', errors='ignore') as nvp_export:
            nvp_export_content = nvp_export.readlines()

        filename = nvpFile.split('\\')[-1]

        nvp_export_traceability[filename] = []

        for line in nvp_export_content:
            match_dom_req = re.match(r'(.*>)(DOM_eCS[^<\n\r]*)', line)

            if match_dom_req is not None:

                temp_dom_str = match_dom_req.group(2).replace(';', ",").replace(' ', '').replace('\t', '')
                temp_dom_str = temp_dom_str.split(',')
                while '' in temp_dom_str:
                    temp_dom_str.remove('')
                for temp in temp_dom_str:
                    nvp_export_traceability[filename].append(temp)

            match_fusa_req = re.match(r'(.*>)(FUSA_[^<\n\r]*)', line)

            if match_fusa_req is not None:

                temp_fusa_str = match_fusa_req.group(2).replace(';', ",").replace(' ', '').replace('\t', '')
                temp_fusa_str = temp_fusa_str.split(',')
                while '' in temp_fusa_str:
                    temp_fusa_str.remove('')
                for temp in temp_fusa_str:
                    nvp_export_traceability[filename].append(temp)

    return nvp_export_traceability


def parse_autosarBuilder_export():

    # Open the text file
    with open("s:\\Tools\\AutosarBuilder\\Workspace\\ABProjects\\SystemConfiguration\\Documents\\SAMPLE-REPORT.html", "r", encoding='utf-8', errors='ignore') as ab_export:
        ab_export_content = ab_export.readlines()

    last_architecture_requirement = ''
    ab_requirement_traceability = {}
    last_port = ''
    dict_with_ARCH_req_and_ports = {}
    list_with_errors_printed = []
    for line in ab_export_content:
        match_port = re.match(r'<td[^>]*><span[^>]*><small><img[^>]*>([^\/]+)<\/small><\/span><\/td>', line)
        match_arch_req_list = re.findall(r'ARCH_SW_[^<>\n\r,; .]*', line)
        match_dom_req_list = re.findall(r'DOM_eCS_[^<>\n\r,; .]*', line)
        match_fusa_req_list = re.findall(r'FUSA_[^<>\n\r,; .]*', line)

        if match_port is not None:
            last_port = match_port.group(1).strip()

        if len(match_arch_req_list) > 0:
            last_architecture_requirement = match_arch_req_list[0]

            if last_architecture_requirement not in dict_with_ARCH_req_and_ports.keys():
                dict_with_ARCH_req_and_ports[last_architecture_requirement] = last_port
                ab_requirement_traceability[last_architecture_requirement] = []
            else:
                if last_port != dict_with_ARCH_req_and_ports[last_architecture_requirement]:
                    errorID = last_architecture_requirement + last_port + dict_with_ARCH_req_and_ports[last_architecture_requirement]
                    if errorID not in list_with_errors_printed:
                        list_with_errors_printed.append(errorID)
                        print('ERROR: Duplicate Arch Req "' + Fore.YELLOW + last_architecture_requirement + Style.RESET_ALL + '" on port "' + Fore.RED + last_port
                          + Style.RESET_ALL + '", already exist on port "' + Fore.GREEN + dict_with_ARCH_req_and_ports[last_architecture_requirement] + Style.RESET_ALL + '"!!!')
                else:
                    dict_with_ARCH_req_and_ports[last_architecture_requirement] = last_port
                    ab_requirement_traceability[last_architecture_requirement] = []

        if len(match_dom_req_list) > 0:
            for match_dom_req in match_dom_req_list:
                ab_requirement_traceability[last_architecture_requirement].append(match_dom_req)

        if len(match_fusa_req_list) > 0:
            for match_fusa_req in match_fusa_req_list:
                ab_requirement_traceability[last_architecture_requirement].append(match_fusa_req)

    return ab_requirement_traceability


# Print iterations progress
def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total:
        print()


def main():
    print("Warning: Make sure AutosarBuilder export is up-to-date")
    print("Warning: Make sure DOORS export is up-to-date")
    print("Warning: Make sure S:/ drive is mounted")

    # Initial call to print 0% progress
    printProgressBar(0, 100, prefix='Progress:', suffix='Complete', length=50)

    # print("Parsing DOORS export")
    srm_statistics = parse_doors_export()

    printProgressBar(12, 100, prefix='Progress:', suffix='Complete', length=50)

    # print("Parsing Integration Test specification")
    int_test_file_list_statistics = parse_integration_tests()

    printProgressBar(24, 100, prefix='Progress:', suffix='Complete', length=50)

    # print("Parsing Unit Test specification")
    unit_test_file_list_statistics = parse_unit_tests()
    MBD_unit_test_list_statistics = parse_MBD_unit_tests()
    unit_test_file_list_statistics.update(MBD_unit_test_list_statistics)

    printProgressBar(36, 100, prefix='Progress:', suffix='Complete', length=50)

    # print("Parsing source code")
    c_file_list_statistics = parse_source_code()

    printProgressBar(48, 100, prefix='Progress:', suffix='Complete', length=50)

    # print("Parsing .xdm files")
    xdm_statistics = parse_xdm_files()

    printProgressBar(60, 100, prefix='Progress:', suffix='Complete', length=50)

    # print("Parsing NVP files")
    nvp_export_traceability = parse_nvp_files()

    printProgressBar(72, 100, prefix='Progress:', suffix='Complete', length=50)

    # print("Parsing AutosarBuilder export")
    ab_requirement_traceability = parse_autosarBuilder_export()

    printProgressBar(84, 100, prefix='Progress:', suffix='Complete', length=50)

    arch_dd_traceability = {}

    warning_print_list = []

    # create DOM <-> XDM traceability
    for key, value in xdm_statistics.items():
        for req in value:
            if req in srm_statistics:
                if srm_statistics[req][2] == 'Obsolete':
                    warning_print_list.append("Hanging DOM requirement found in xdm files: " + req)
                else:
                    if key not in srm_statistics[req]:
                        srm_statistics[req].append(key)
            else:
                warning_print_list.append("Hanging DOM requirement found in xdm files: " + req)

    # create DOM <-> NVP traceability
    for key, value in nvp_export_traceability.items():
        for req in value:
            if req in srm_statistics:
                if srm_statistics[req][2] == 'Obsolete':
                    warning_print_list.append("Hanging DOM requirement found in NVP files: " + req)
                else:
                    if key not in srm_statistics[req]:
                        srm_statistics[req].append(key)
            else:
                warning_print_list.append("Hanging DOM requirement found in NVP files: " + req)

    # convert dict[DSG] = [ARCHs] to dict[ARCH] = [DSGs]
    dict_DSG_traceability_in_ARCH = {}
    for DSG_req, ARCH_reqs_list in c_file_list_statistics.items():
        for ARCH_req in ARCH_reqs_list:
            if ARCH_req not in dict_DSG_traceability_in_ARCH.keys():
                dict_DSG_traceability_in_ARCH[ARCH_req] = [DSG_req]
            else:
                if DSG_req not in dict_DSG_traceability_in_ARCH[ARCH_req]:
                    dict_DSG_traceability_in_ARCH[ARCH_req].append(DSG_req)
    for ARCH_req in ab_requirement_traceability.keys():
        if ARCH_req not in dict_DSG_traceability_in_ARCH.keys():
            dict_DSG_traceability_in_ARCH[ARCH_req] = []
            print('Architecture requirement "' + ARCH_req + '" does not have traceability in Detailed Design!')

    # create DOM <-> ARCH traceability
    # create DD <-> ARCH traceability
    for key, value in ab_requirement_traceability.items():
        for req in value:
            if req in srm_statistics:
                if srm_statistics[req][2] == 'Obsolete':
                    warning_print_list.append("Hanging DOM requirement found in AutosarBuilder export: " + req)
                else:
                    srm_statistics[req].append(key)
                    for key2, value2 in c_file_list_statistics.items():
                        for req2 in value2:
                            if req2 == key:
                                arch_dd_traceability[req2] = [key2]
            else:
                warning_print_list.append("Hanging DOM requirement found in AutosarBuilder export: " + req)

    # create IT <-> ARCH traceability
    arch_it_traceability = {}
    for key, value in ab_requirement_traceability.items():
        for req in value:
            if req in srm_statistics.keys():
                srm_statistics[req].append(key)
                for key2, value2 in int_test_file_list_statistics.items():
                    if key in value2:
                        arch_it_traceability[req2] = [os.path.basename(key2)]
                    else:
                        warning_print_list.append("Hanging ARCH_SW requirement found in IT files: " + key)

    # convert dict[IT] = [ARCHs] to dict[ARCH] = [ITs]
    dict_ARCH_traceability_in_IT = {}
    for IT_name, ARCH_reqs_list in int_test_file_list_statistics.items():
        for ARCH_req in ARCH_reqs_list:
            if ARCH_req not in dict_ARCH_traceability_in_IT.keys():
                dict_ARCH_traceability_in_IT[ARCH_req] = [IT_name]
            else:
                if IT_name not in dict_ARCH_traceability_in_IT[ARCH_req]:
                    dict_ARCH_traceability_in_IT[ARCH_req].append(IT_name)
    for ARCH_req in ab_requirement_traceability.keys():
        if ARCH_req not in dict_ARCH_traceability_in_IT.keys():
            dict_ARCH_traceability_in_IT[ARCH_req] = []
            print('Architecture requirement "' + ARCH_req + '" does not have traceability in Integration Tests!')

    # create DD <-> UT traceability
    dd_ut_tracebility = {}
    for key, value in c_file_list_statistics.items():
        for req in value:
            for key3, value3 in srm_statistics.items():
                if req in value3:
                    for key2, value2 in unit_test_file_list_statistics.items():
                        if key in value2:
                            dd_ut_tracebility[req2] = [os.path.basename(key2)]
                        else:
                            warning_print_list.append("Hanging DDSG requirement found in UT files: " + key)

    # convert dict[UT] = [DSGs] to dict[DSG] = [UTs]
    dict_DSG_traceability_in_UT = {}
    for UT_name, DSG_reqs_list in unit_test_file_list_statistics.items():
        for DSG_req in DSG_reqs_list:
            if DSG_req not in dict_DSG_traceability_in_UT.keys():
                dict_DSG_traceability_in_UT[DSG_req] = [UT_name]
            else:
                if UT_name not in dict_DSG_traceability_in_UT[DSG_req]:
                    dict_DSG_traceability_in_UT[DSG_req].append(UT_name)
    for DSG_req in c_file_list_statistics.keys():
        if DSG_req not in dict_DSG_traceability_in_UT.keys():
            dict_DSG_traceability_in_UT[DSG_req] = []
            print('Detailed Design requirement "' + DSG_req + '" does not have traceability in Unit Tests!')

    printProgressBar(90, 100, prefix='Progress:', suffix='Complete', length=50)

    # Create a new Excel workbook
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet("Overview", cell_overwrite_ok=True)
    xlwt.add_palette_colour("custom_colour", 0x21)
    workbook.set_colour_RGB(0x21, 255, 102, 102)
    style = xlwt.easyxf('pattern: pattern solid, fore_colour custom_colour')

    xlwt.add_palette_colour("custom_colour1", 0x22)
    workbook.set_colour_RGB(0x22, 255, 255, 204)
    style2 = xlwt.easyxf('pattern: pattern solid, fore_colour custom_colour1')

    xlwt.add_palette_colour("custom_colour2", 0x23)
    workbook.set_colour_RGB(0x23, 204, 255, 204)
    style3 = xlwt.easyxf('pattern: pattern solid, fore_colour custom_colour2')

    # Write each match to a separate cell in the Excel worksheet
    worksheet.write(0, 0, '_ID')
    worksheet.write(0, 1, 'Detailed_Design_Traceability')
    worksheet.write(0, 2, 'IntegrationTest_Traceability')
    worksheet.write(0, 3, 'UnitTest_Traceability')
    worksheet.write(0, 4, '_Target_eCS')
    worksheet.write(0, 5, '_Implemented')
    worksheet.write(0, 6, '_ReqStatus')
    worksheet.write(0, 7, '_ImpactedDiscipline')
    worksheet.write(0, 8, '_ASIL')
    worksheet.write(0, 9, 'Architecture_Traceability')
    linenb = 1
    for key, value in srm_statistics.items():

        if(len(srm_statistics[key]) <= 5):
            worksheet.write(linenb, 0, key, style)
        else:
            worksheet.write(linenb, 0, key)

        mylinenb = 0

        for idx, elem in enumerate(list(dict.fromkeys(value))):
            if mylinenb == 1 and elem == 'False':
                worksheet.write(linenb, 4 + mylinenb, elem.replace(',', ""), style2)
            else:
                worksheet.write(linenb, 4 + mylinenb, elem.replace(',', ""))
                if idx == 5 and ('.xdm' in elem or '.xml' in elem):
                    worksheet.write(linenb, 1, 'N/A.', style3)
                    worksheet.write(linenb, 2, 'N/A.', style3)
                    worksheet.write(linenb, 3, 'N/A.', style3)
                if "ARCH_SW" in elem:
                    if len(dict_DSG_traceability_in_ARCH[elem]) > 0:
                        dsg_requirement = dict_DSG_traceability_in_ARCH[elem][0]
                        worksheet.write(linenb, 1, dsg_requirement, style3)
                        for dsgReq in dict_DSG_traceability_in_ARCH[elem]:
                            if dsgReq in dict_DSG_traceability_in_UT:
                                if len(dict_DSG_traceability_in_UT[dsgReq]) > 0:
                                    worksheet.write(linenb, 3, dict_DSG_traceability_in_UT[dsgReq][0], style3)
                                    break;
                    if elem in dict_ARCH_traceability_in_IT.keys():
                        if len(dict_ARCH_traceability_in_IT[elem]) > 0:
                            worksheet.write(linenb, 2, dict_ARCH_traceability_in_IT[elem][0], style3)

            mylinenb += 1

        linenb += 1

    printProgressBar(100, 100, prefix='Progress:', suffix='Complete', length=50)

    for warning in set(warning_print_list):
        print(warning)

    worksheet_Arch = workbook.add_sheet("_SwDetailedDesign", cell_overwrite_ok=True)
    worksheet_Int = workbook.add_sheet("_SwIntegrationTest", cell_overwrite_ok=True)
    worksheet_Ut = workbook.add_sheet("_SwUnitTest", cell_overwrite_ok=True)
    worksheet_ARCH_IT = workbook.add_sheet("_SwIntegrationTestInARCH", cell_overwrite_ok=True)
    worksheet_DSG_UT = workbook.add_sheet("_SwUnitTestInDSG", cell_overwrite_ok=True)
    worksheet_DD = workbook.add_sheet("_SwArchitecture", cell_overwrite_ok=True)


    worksheet_Int.write(0, 0, '_IntTest')
    worksheet_Int.write(0, 1, '_SwArch_Traceability')

    linenb = 1
    for key, value in int_test_file_list_statistics.items():
        worksheet_Int.write(linenb, 0, key)
        mylinenb = 0
        for req in value:
            worksheet_Int.write(linenb, 1+mylinenb, req)
            mylinenb += 1
        linenb += 1


    worksheet_ARCH_IT.write(0, 0, '_SwArch')
    worksheet_ARCH_IT.write(0, 1, '_IntTest_Traceability')

    linenb = 1
    for key, value in dict_ARCH_traceability_in_IT.items():
        worksheet_ARCH_IT.write(linenb, 0, key)
        mylinenb = 0
        for req in value:
            worksheet_ARCH_IT.write(linenb, 1 + mylinenb, req)
            mylinenb += 1
        linenb += 1


    worksheet_Arch.write(0, 0, '_DSG_Req')
    worksheet_Arch.write(0, 1, '_SwArch_Traceability')

    linenb = 1
    for key, value in c_file_list_statistics.items():
        worksheet_Arch.write(linenb, 0, key)
        mylinenb = 0
        for req in value:
            worksheet_Arch.write(linenb, 1 + mylinenb, req)
            mylinenb += 1
        linenb += 1


    worksheet_Ut.write(0, 0, '_UnitTest')
    worksheet_Ut.write(0, 1, '_SwDDsg_Traceability')

    linenb = 1
    for key, value in unit_test_file_list_statistics.items():
        worksheet_Ut.write(linenb, 0, key)
        mylinenb = 0
        for req in value:
            worksheet_Ut.write(linenb, 1 + mylinenb, req)
            mylinenb += 1
        linenb += 1


    worksheet_DSG_UT.write(0, 0, '_SwDDsg')
    worksheet_DSG_UT.write(0, 1, '_UnitTest_Traceability')

    linenb = 1
    for key, value in dict_DSG_traceability_in_UT.items():
        worksheet_DSG_UT.write(linenb, 0, key)
        mylinenb = 0
        for req in value:
            worksheet_DSG_UT.write(linenb, 1 + mylinenb, req)
            mylinenb += 1
        linenb += 1


    worksheet_DD.write(0, 0, '_SwArchReq')
    worksheet_DD.write(0, 1, '_SwDDsg_Traceability')

    linenb = 1
    for key, value in dict_DSG_traceability_in_ARCH.items():
        worksheet_DD.write(linenb, 0, key)
        mylinenb = 0
        for req in value:
            worksheet_DD.write(linenb, 1 + mylinenb, req)
            mylinenb += 1
        linenb += 1
    print("All done. Live long and prosper")

    # Save the Excel workbook
    workbook.save("SwTraceabilityMatrix.xls")

if __name__ == '__main__':
    main()
