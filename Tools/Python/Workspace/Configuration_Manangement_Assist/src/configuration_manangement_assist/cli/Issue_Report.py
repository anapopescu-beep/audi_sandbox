"""Issue_Report.py: Used to generate an issue report from excel API of OpenProject

"""

__author__ = 'Emilian Gustescu'
__copyright__ = "Copyright 2022, Autoliv Electronic"
__version__ = "1.0.0"
__email__ = 'emilian.gustescu@autoliv.com'
__status__ = 'Released'


import win32com.client as win32
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import argparse
import os


def main():

    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input-path-excel-template', help="Path to .xlsm excel template file. Path must also contain the filename and extension.", required=True)
    parser.add_argument('-t', '--input-path-excel-library', help="Path to .xlam excel library file. Path must also contain the filename and extension", required=True)
    parser.add_argument('-o', '--output-file', help="Path to output Excel file.", required=True)
    args = parser.parse_args()

    path_to_excel_template = args.input_path_excel_template
    path_to_excel_library = args.input_path_excel_library
    path_to_output_file = args.output_file

    # task kill excel
    os.system('Taskkill /IM EXCEL.EXE /F')

    print("Running script")
    print("==========================================")
    # Start Excel application
    excel1 = win32.Dispatch("Excel.Application")
    excel2 = win32.Dispatch("Excel.Application")

    # Get the path of the Excel library and Excel template file
    path1 = path_to_excel_library
    path2 = path_to_excel_template

    # Open the Excel library and Excel template
    wb1 = excel1.Workbooks.Open(path1)
    wb2 = excel2.Workbooks.Open(path2)

    # Hide both Excel files
    excel1.Visible = False
    excel2.Visible = False

    # Activate only the worksheet where we will collect our data
    excel2.Worksheets("OpenProjectData").Activate()

    # Use the button to clear old data and collect the one from OpenProject
    try:
        print("Fetching Data...")
        excel2.Application.Run("API_ClearAndLoadAll")
    except:
        print("Failed to load data")
        pass

    # Save the data to the Excel template
    excel2.ActiveWorkbook.Save()

    os.system('Taskkill /IM EXCEL.EXE /F')

    wb = load_workbook(path_to_excel_template, read_only=False)
    warnings.simplefilter(action='ignore', category=UserWarning)

    sheet = wb["OpenProjectData"]

    id_list = []
    issue_type_list = []
    state_list = []
    created_list = []
    assigned_list = []
    title_list = []
    current_row = 5
    for row in sheet.iter_rows(min_row=6, min_col=1, max_row=sheet.max_row):
        current_row += 1
        if sheet.cell(current_row, 2).value != None:
            id_list.append(sheet.cell(current_row, 2).value)
            issue_type_list.append(sheet.cell(current_row, 4).value)
            state_list.append(sheet.cell(current_row, 6).value)
            created_list.append(sheet.cell(current_row, 11).value)
            assigned_list.append(sheet.cell(current_row, 12).value)
            title_list.append(sheet.cell(current_row, 9).value)


    print("==========================================")
    print("Data fetched")
    print("==========================================")
    print("Processing data...")
    print("==========================================")

    wb = load_workbook(path_to_output_file, read_only=False)
    warnings.simplefilter(action='ignore', category=UserWarning)

    sheet = wb["Issue_Report"]

    sheet['A1'] = "ID"
    sheet['A1'].font = Font(bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet['A1'].fill = PatternFill("solid", start_color="e60000")
    sheet.column_dimensions['A'].width = 15

    sheet['B1'] = "Issue Type"
    sheet['B1'].font = Font(bold=True)
    sheet['B1'].alignment = Alignment(horizontal='center')
    sheet['B1'].fill = PatternFill("solid", start_color="e60000")
    sheet.column_dimensions['B'].width = 15

    sheet['C1'] = "State"
    sheet['C1'].font = Font(bold=True)
    sheet['C1'].alignment = Alignment(horizontal='center')
    sheet['C1'].fill = PatternFill("solid", start_color="e60000")
    sheet.column_dimensions['C'].width = 35

    sheet['D1'] = "Currently Assigned To"
    sheet['D1'].font = Font(bold=True)
    sheet['D1'].alignment = Alignment(horizontal='center')
    sheet['D1'].fill = PatternFill("solid", start_color="e60000")
    sheet.column_dimensions['D'].width = 20

    sheet['E1'] = "Created by"
    sheet['E1'].font = Font(bold=True)
    sheet['E1'].alignment = Alignment(horizontal='center')
    sheet['E1'].fill = PatternFill("solid", start_color="e60000")
    sheet.column_dimensions['E'].width = 20

    sheet['F1'] = "Issue Title"
    sheet['F1'].font = Font(bold=True)
    sheet['F1'].alignment = Alignment(horizontal='center')
    sheet['F1'].fill = PatternFill("solid", start_color="e60000")
    sheet.column_dimensions['F'].width = 100

    sheet['G1'] = "Verified on Software Version"
    sheet['G1'].font = Font(bold=True)
    sheet['G1'].alignment = Alignment(horizontal='center')
    sheet['G1'].fill = PatternFill("solid", start_color="e60000")
    sheet.column_dimensions['G'].width = 35

    sheet['H1'] = "Verification result"
    sheet['H1'].font = Font(bold=True)
    sheet['H1'].alignment = Alignment(horizontal='center')
    sheet['H1'].fill = PatternFill("solid", start_color="e60000")
    sheet.column_dimensions['H'].width = 35

    sheet['I1'] = "Verification Assigned"
    sheet['I1'].font = Font(bold=True)
    sheet['I1'].alignment = Alignment(horizontal='center')
    sheet['I1'].fill = PatternFill("solid", start_color="e60000")
    sheet.column_dimensions['I'].width = 35

    wb.save(path_to_output_file)

    wb = load_workbook(path_to_output_file, read_only=False)
    warnings.simplefilter(action='ignore', category=UserWarning)

    sheet = wb["Issue_Report"]

    # Add new issues
    current_element = 0
    for elem in id_list:
        in_file = False
        current_row = 1
        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=len(id_list)+1):
            current_row+=1
            if sheet.cell(current_row, 1).value == elem:
                in_file = True
        if in_file == False:
            current_inner_row = 1
            for row in sheet.iter_rows(min_row=2, min_col=1, max_row=len(id_list) + 1):
                current_inner_row += 1
                if sheet.cell(current_inner_row, 1).value == None:
                    sheet.cell(current_inner_row, 1).value = id_list[current_element]
                    sheet.cell(current_inner_row, 2).value = issue_type_list[current_element]
                    sheet.cell(current_inner_row, 3).value = state_list[current_element]
                    sheet.cell(current_inner_row, 4).value = created_list[current_element]
                    sheet.cell(current_inner_row, 5).value = assigned_list[current_element]
                    sheet.cell(current_inner_row, 6).value = title_list[current_element]
                    break

        current_element+=1

    # Check existing issues for changes
    current_row = 1
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row):
        current_row += 1
        current_element = 0
        for elem in id_list:
            if sheet.cell(current_row, 1).value == elem:
                if sheet.cell(current_row, 2).value != issue_type_list[current_element]:
                    sheet.cell(current_row, 2).value = issue_type_list[current_element]
                if sheet.cell(current_row, 3).value != state_list[current_element]:
                    sheet.cell(current_row, 3).value = state_list[current_element]
                if sheet.cell(current_row, 4).value != created_list[current_element]:
                    sheet.cell(current_row, 4).value = created_list[current_element]
                if sheet.cell(current_row, 5).value != assigned_list[current_element]:
                    sheet.cell(current_row, 5).value = assigned_list[current_element]
                if sheet.cell(current_row, 6).value != title_list[current_element]:
                    sheet.cell(current_row, 6).value = title_list[current_element]
            current_element+=1

    current_row = 1
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row):
        current_row += 1
        if current_row % 2 == 0:
            sheet.cell(current_row, 1).fill = PatternFill("solid", start_color="d1d1e0")
            sheet.cell(current_row, 2).fill = PatternFill("solid", start_color="d1d1e0")
            sheet.cell(current_row, 3).fill = PatternFill("solid", start_color="d1d1e0")
            sheet.cell(current_row, 4).fill = PatternFill("solid", start_color="d1d1e0")
            sheet.cell(current_row, 5).fill = PatternFill("solid", start_color="d1d1e0")
            sheet.cell(current_row, 6).fill = PatternFill("solid", start_color="d1d1e0")
            sheet.cell(current_row, 7).fill = PatternFill("solid", start_color="d1d1e0")
            sheet.cell(current_row, 8).fill = PatternFill("solid", start_color="d1d1e0")
            sheet.cell(current_row, 9).fill = PatternFill("solid", start_color="d1d1e0")

    wb.save(path_to_output_file)

    # excel2.Quit()
    # try:
    #     excel1.Quit()
    # except:
    #     pass

    print("Done processing data")
    print("==========================================")
    print("Script ran successfully")
    print("==========================================")

if __name__ == "__main__":
    main()
