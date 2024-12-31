"""Process_OpenProject_Data.py: Used to generate KPI statistics from OpenProject
"""

__author__ = 'Emilian Gustescu'
__copyright__ = "Copyright 2022, Autoliv Electronic"
__version__ = "1.0.0"
__email__ = 'emilian.gustescu@autoliv.com'
__status__ = 'Released'

import win32com.client as win32
import warnings
from openpyxl import load_workbook
from matplotlib import pyplot as plt
import pandas as pd
import numpy as np
from textwrap import wrap
from datetime import datetime
import argparse
import os
import csv

def story_task_data(path_to_excel_template, path_to_output_folder):

    # Load template Excel file and pick the OpenProjectData sheet
    wb = load_workbook(path_to_excel_template, read_only=True)
    warnings.simplefilter(action='ignore', category=UserWarning)

    sheet = wb["OpenProjectData"]

    # Create dictionaries for stories and tasks to count their statuses
    storydict = {
        "New": 0,
        "In progress": 0,
        "On hold": 0,
        "Closed": 0
    }
    taskdict = {
        "New": 0,
        "In progress": 0,
        "On hold": 0,
        "Closed": 0
    }

    # Iterate through Excel file and check for every Story and Task their status and count it
    current_row = 5
    for row in sheet.iter_rows(min_row=6, min_col=4, max_row=sheet.max_row, max_col=14):
        current_row = current_row + 1
        for cell in row:
            if cell.value == 'Story':
                if sheet.cell(current_row, 6).value == 'New':
                    storydict["New"] = storydict["New"] + 1
                elif sheet.cell(current_row, 6).value == 'In progress':
                    storydict["In progress"] = storydict["In progress"] + 1
                elif sheet.cell(current_row, 6).value == 'On hold':
                    storydict["On hold"] = storydict["On hold"] + 1
                elif sheet.cell(current_row, 6).value == 'Closed':
                    storydict["Closed"] = storydict["Closed"] + 1
            elif cell.value == 'AEM Task':
                if sheet.cell(current_row, 6).value == 'New':
                    taskdict["New"] = taskdict["New"] + 1
                elif sheet.cell(current_row, 6).value == 'In progress':
                    taskdict["In progress"] = taskdict["In progress"] + 1
                elif sheet.cell(current_row, 6).value == 'On hold':
                    taskdict["On hold"] = taskdict["On hold"] + 1
                elif sheet.cell(current_row, 6).value == 'Closed':
                    taskdict["Closed"] = taskdict["Closed"] + 1

    # Add dictionaries to chart data
    storydata = [storydict["New"], storydict["In progress"], storydict["On hold"], storydict["Closed"]]
    taskdata = [taskdict["New"], taskdict["In progress"], taskdict["On hold"], taskdict["Closed"]]

    # Create chart with the status on X axis
    X =['New', 'In progress', 'On hold', 'Closed']
    X_axis = np.arange(len(X))
    storybars = plt.bar(X_axis - 0.2, storydata, 0.4, label='Story')
    taskbars = plt.bar(X_axis + 0.2, taskdata, 0.4, label='Task')

    plt.xticks(X_axis, X)
    plt.xlabel("Status")
    plt.ylabel("Count")
    plt.title("Count status for Stories and Tasks")
    plt.legend(loc='center left', bbox_to_anchor=(1,0.5))
    plt.tight_layout()

    # Add numbers on each bar for data to be easier to read
    bar_color = storybars[0].get_facecolor()
    for bar in storybars:
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.01,
            round(bar.get_height(), 1),
            horizontalalignment='center',
            color=bar_color,
            weight='bold'
        )

    bar_color = taskbars[0].get_facecolor()
    for bar in taskbars:
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.01,
            round(bar.get_height(), 1),
            horizontalalignment='center',
            color=bar_color,
            weight='bold'
        )

    plt.savefig(path_to_output_folder + "\status_story_task.png")
    plt.close()


def bug_feature_data(path_to_excel_template, path_to_output_folder):

    # Load template Excel file and pick the OpenProjectData sheet
    path = path_to_excel_template
    wb = load_workbook(path, read_only=True)
    warnings.simplefilter(action='ignore', category=UserWarning)

    sheet = wb["OpenProjectData"]

    # Create dictionaries for bugs and features to count their statuses
    bugdict = {
        "New": 0,
        "Closed": 0,
        "To Be Analysed": 0,
        "Analysis On Going": 0,
        "Analysed": 0,
        "To Be Implemented": 0,
        "Implementation On Going": 0,
        "Implementation To Be Reviewed": 0,
        "Implementation Review On Going": 0,
        "Postponed": 0
    }

    featuredict = {
        "New": 0,
        "Closed": 0,
        "To Be Analysed": 0,
        "Analysis On Going": 0,
        "Analysed": 0,
        "To Be Implemented": 0,
        "Implementation On Going": 0,
        "Implementation To Be Reviewed": 0,
        "Implementation Review On Going": 0,
        "Postponed": 0
    }

    # Iterate through Excel file and check for every Bug and Feature their status and count it
    current_row = 5
    for row in sheet.iter_rows(min_row=6, min_col=4, max_row=sheet.max_row, max_col=14):
        current_row = current_row + 1
        for cell in row:
            if cell.value == 'Bug':
                if sheet.cell(current_row, 6).value == 'New':
                    bugdict["New"] = bugdict["New"] + 1
                elif sheet.cell(current_row, 6).value == 'Closed':
                    bugdict["Closed"] = bugdict["Closed"] + 1
                elif sheet.cell(current_row, 6).value == 'To Be Analysed':
                    bugdict["To Be Analysed"] = bugdict["To Be Analysed"] + 1
                elif sheet.cell(current_row, 6).value == 'Analysis On Going':
                    bugdict["Analysis On Going"] = bugdict["Analysis On Going"] + 1
                elif sheet.cell(current_row, 6).value == 'Analysed':
                    bugdict["Analysed"] = bugdict["Analysed"] + 1
                elif sheet.cell(current_row, 6).value == 'To Be Implemented':
                    bugdict["To Be Implemented"] = bugdict["To Be Implemented"] + 1
                elif sheet.cell(current_row, 6).value == 'Implementation On Going':
                    bugdict["Implementation On Going"] = bugdict["Implementation On Going"] + 1
                elif sheet.cell(current_row, 6).value == 'Implementation To Be Reviewed':
                    bugdict["Implementation To Be Reviewed"] = bugdict["Implementation To Be Reviewed"] + 1
                elif sheet.cell(current_row, 6).value == 'Implementation Review On Going':
                    bugdict["Implementation Review On Going"] = bugdict["Implementation Review On Going"] + 1
                elif sheet.cell(current_row, 6).value == 'Postponed':
                    bugdict["Postponed"] = bugdict["Postponed"] + 1

            if cell.value == 'Feature':
                if sheet.cell(current_row, 6).value == 'New':
                    featuredict["New"] = featuredict["New"] + 1
                elif sheet.cell(current_row, 6).value == 'Closed':
                    featuredict["Closed"] = featuredict["Closed"] + 1
                elif sheet.cell(current_row, 6).value == 'To Be Analysed':
                    featuredict["To Be Analysed"] = featuredict["To Be Analysed"] + 1
                elif sheet.cell(current_row, 6).value == 'Analysis On Going':
                    featuredict["Analysis On Going"] = featuredict["Analysis On Going"] + 1
                elif sheet.cell(current_row, 6).value == 'Analysed':
                    featuredict["Analysed"] = featuredict["Analysed"] + 1
                elif sheet.cell(current_row, 6).value == 'To Be Implemented':
                    featuredict["To Be Implemented"] = featuredict["To Be Implemented"] + 1
                elif sheet.cell(current_row, 6).value == 'Implementation On Going':
                    featuredict["Implementation On Going"] = featuredict["Implementation On Going"] + 1
                elif sheet.cell(current_row, 6).value == 'Implementation To Be Reviewed':
                    featuredict["Implementation To Be Reviewed"] = featuredict["Implementation To Be Reviewed"] + 1
                elif sheet.cell(current_row, 6).value == 'Implementation Review On Going':
                    featuredict["Implementation Review On Going"] = featuredict["Implementation Review On Going"] + 1
                elif sheet.cell(current_row, 6).value == 'Postponed':
                    featuredict["Postponed"] = featuredict["Postponed"] + 1

    # Add dictionaries to chart data
    featuredata = [featuredict["New"],
                 featuredict["Closed"],
                 featuredict["To Be Analysed"],
                 featuredict["Analysis On Going"],
                 featuredict["Analysed"],
                 featuredict["To Be Implemented"],
                 featuredict["Implementation On Going"],
                 featuredict["Implementation To Be Reviewed"],
                 featuredict["Implementation Review On Going"],
                 featuredict["Postponed"]]

    bugdata = [bugdict["New"],
                bugdict["Closed"],
                bugdict["To Be Analysed"],
                bugdict["Analysis On Going"],
                bugdict["Analysed"],
                bugdict["To Be Implemented"],
                bugdict["Implementation On Going"],
                bugdict["Implementation To Be Reviewed"],
                bugdict["Implementation Review On Going"],
                bugdict["Postponed"]]

    # Create chart with the status on X axis
    X = ['New', 'Closed', 'To Be Analysed', 'Analysis On Going', 'Analysed', 'To Be Implemented', 'Implementation On Going', 'Implementation To Be Reviewed', 'Implementation Review On Going', 'Postponed']
    X_axis = np.arange(len(X))
    plt.figure(figsize=(30, 10))
    featurebars = plt.bar(x=X_axis - 0.2, height=featuredata, width=0.4, label='Feature')
    bugbars = plt.bar(x=X_axis + 0.2, height=bugdata, width=0.4, data=bugdata, label='Bug')

    plt.xticks(X_axis, X, rotation=15)
    plt.xlabel("Status")
    plt.ylabel("Count")
    plt.title("Count status for Bugs and Features")

    # Add numbers on each bar for data to be easier to read
    plt.legend(loc='center left', bbox_to_anchor=(1,0.5))
    plt.tight_layout()
    bar_color = featurebars[0].get_facecolor()
    for bar in featurebars:
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.01,
            round(bar.get_height(), 1),
            horizontalalignment='center',
            color=bar_color,
            weight='bold'
        )

    bar_color = bugbars[0].get_facecolor()
    for bar in bugbars:
        plt.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.01,
            round(bar.get_height(), 1),
            horizontalalignment='center',
            color=bar_color,
            weight='bold'
        )

    plt.savefig(path_to_output_folder + '\status_feature_bug.png')
    plt.close()


def user_data(path_to_excel_template, path_to_output_folder):

    # Load template Excel file and pick the OpenProjectData sheet
    path = path_to_excel_template
    wb = load_workbook(path, read_only=True)
    warnings.simplefilter(action='ignore', category=UserWarning)

    sheet = wb["OpenProjectData"]

    # Add chart data to a dictionary of dictionaries
    userdict = {}
    for row in sheet.iter_rows(min_row=6, min_col=11, max_row=sheet.max_row, max_col=11):
        for cell in row:
            userdict.update({cell.value: {
                "New": 0,
                "In progress": 0,
                "On hold": 0,
                "Closed": 0,
                "To Be Analysed": 0,
                "Analysis On Going": 0,
                "Analysed": 0,
                "To Be Implemented": 0,
                "Implementation On Going": 0,
                "Implementation To Be Reviewed": 0,
                "Implementation Review On Going": 0,
                "Postponed": 0
            }})

    # Remove all fields that are "None"
    index = ""
    for i in userdict.keys():
        index = i
    if index is None:
        userdict.popitem()

    # Iterate through Excel file and check for every user how many issues they have and their status
    current_row = 5
    for row in sheet.iter_rows(min_row=6, min_col=4, max_row=sheet.max_row, max_col=14):
        current_row = current_row + 1
        for cell in row:
            for userName in userdict:
                if cell.value == userName and cell.value is not None:
                    user = userdict[userName]
                    if sheet.cell(current_row, 6).value == 'New':
                        user["New"] = user["New"] + 1
                    elif sheet.cell(current_row, 6).value == 'In progress':
                        user["In progress"] = user["In progress"] + 1
                    elif sheet.cell(current_row, 6).value == 'On hold':
                        user["On hold"] = user["On hold"] + 1
                    elif sheet.cell(current_row, 6).value == 'Closed':
                        user["Closed"] = user["Closed"] + 1
                    elif sheet.cell(current_row, 6).value == 'To Be Analysed':
                        user["To Be Analysed"] = user["To Be Analysed"] + 1
                    elif sheet.cell(current_row, 6).value == 'Analysis On Going':
                        user["Analysis On Going"] = user["Analysis On Going"] + 1
                    elif sheet.cell(current_row, 6).value == 'Analysed':
                        user["Analysed"] = user["Analysed"] + 1
                    elif sheet.cell(current_row, 6).value == 'To Be Implemented':
                        user["To Be Implemented"] = user["To Be Implemented"] + 1
                    elif sheet.cell(current_row, 6).value == 'Implementation On Going':
                        user["Implementation On Going"] = user["Implementation On Going"] + 1
                    elif sheet.cell(current_row, 6).value == 'Implementation To Be Reviewed':
                        user["Implementation To Be Reviewed"] = user["Implementation To Be Reviewed"] + 1
                    elif sheet.cell(current_row, 6).value == 'Implementation Review On Going':
                        user["Implementation Review On Going"] = user["Implementation Review On Going"] + 1
                    elif sheet.cell(current_row, 6).value == 'Postponed':
                        user["Postponed"] = user["Postponed"] + 1

    plt.figure(figsize=(10, 7))

    # Save data from user dictionary to be used in the chart
    user_names = []
    state_dict = {}
    for userName in userdict:
        user = userdict[userName]
        user_names.append(userName)
        for state in user:
            try:
                state_dict[state].append(user[state])
            except:
                state_dict[state] = [user[state]]

    # Create bar chart
    plotdata = pd.DataFrame(state_dict, user_names)
    X_axis = np.arange(len(user_names))
    plotdata.plot(kind="bar", stacked=True, figsize=(10, 10))
    plt.xticks(X_axis, rotation=25)
    plt.title("Count state for each user")
    plt.xlabel("Users")
    plt.ylabel("State count")
    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
    plt.tight_layout()

    plt.savefig(path_to_output_folder + '\status_user.png')
    plt.close()


def sprint_status_data(path_to_excel_template, path_to_output_folder):

    # Load template Excel file and pick the OpenProjectData sheet
    path = path_to_excel_template
    wb = load_workbook(path, read_only=True)
    warnings.simplefilter(action='ignore', category=UserWarning)

    sheet = wb["OpenProjectData"]

    # Add data to a dictionary of dictionaries for easier manipulation
    sprintdict = {}
    for row in sheet.iter_rows(min_row=6, min_col=15, max_row=sheet.max_row, max_col=15):
        for cell in row:
            if cell.value is not None:
                sprintdict.update({cell.value: {
                    "New": 0,
                    "In progress": 0,
                    "On hold": 0,
                    "Closed": 0,
                    "To Be Analysed": 0,
                    "Analysis On Going": 0,
                    "Analysed": 0,
                    "To Be Implemented": 0,
                    "Implementation On Going": 0,
                    "Implementation To Be Reviewed": 0,
                    "Implementation Review On Going": 0,
                    "Postponed": 0
                }})

    current_row = 5

    # Iterate through Excel file and check for every sprint their issues and then count them
    for row in sheet.iter_rows(min_row=6, min_col=4, max_row=sheet.max_row, max_col=15):
        current_row = current_row + 1
        for cell in row:
            for sprintName in sprintdict:
                if cell.value == sprintName and cell.value is not None:
                    sprint = sprintdict[sprintName]
                    if sheet.cell(current_row, 6).value == 'New':
                        sprint["New"] = sprint["New"] + 1
                    elif sheet.cell(current_row, 6).value == 'In progress':
                        sprint["In progress"] = sprint["In progress"] + 1
                    elif sheet.cell(current_row, 6).value == 'On hold':
                        sprint["On hold"] = sprint["On hold"] + 1
                    elif sheet.cell(current_row, 6).value == 'Closed':
                        sprint["Closed"] = sprint["Closed"] + 1
                    elif sheet.cell(current_row, 6).value == 'To Be Analysed':
                        sprint["To Be Analysed"] = sprint["To Be Analysed"] + 1
                    elif sheet.cell(current_row, 6).value == 'Analysis On Going':
                        sprint["Analysis On Going"] = sprint["Analysis On Going"] + 1
                    elif sheet.cell(current_row, 6).value == 'Analysed':
                        sprint["Analysed"] = sprint["Analysed"] + 1
                    elif sheet.cell(current_row, 6).value == 'To Be Implemented':
                        sprint["To Be Implemented"] = sprint["To Be Implemented"] + 1
                    elif sheet.cell(current_row, 6).value == 'Implementation On Going':
                        sprint["Implementation On Going"] = sprint["Implementation On Going"] + 1
                    elif sheet.cell(current_row, 6).value == 'Implementation To Be Reviewed':
                        sprint["Implementation To Be Reviewed"] = sprint["Implementation To Be Reviewed"] + 1
                    elif sheet.cell(current_row, 6).value == 'Implementation Review On Going':
                        sprint["Implementation Review On Going"] = sprint["Implementation Review On Going"] + 1
                    elif sheet.cell(current_row, 6).value == 'Postponed':
                        sprint["Postponed"] = sprint["Postponed"] + 1

    # Save data from dictionary in a list to use in the chart
    sprint_names = []
    state_dict = {}
    for sprintName in sprintdict:
        sprint = sprintdict[sprintName]
        sprint_names.append(sprintName)
        for state in sprint:
            try:
                state_dict[state].append(sprint[state])
            except:
                state_dict[state] = [sprint[state]]

    # Use data saved from dictionaries to plot bar chart
    plotdata = pd.DataFrame(state_dict, sprint_names)
    X_axis = np.arange(len(sprint_names))
    plotdata.plot(kind="bar", stacked=True, figsize=(20, 15))
    labels = [ '\n'.join(wrap(l, 20)) for l in sprint_names]
    plt.rcParams.update({'font.size': 15})
    plt.xticks(X_axis, labels, rotation=90)
    plt.title("Count state for each sprint")
    plt.xlabel("Sprint")
    plt.ylabel("State count")
    plt.legend(loc='center left', bbox_to_anchor=(1,0.5))
    plt.tight_layout()

    plt.savefig(path_to_output_folder + '\status_sprint.png')
    plt.close()


def sprint_type_data(path_to_excel_template, path_to_output_folder):

    # Load template Excel file and pick the OpenProjectData sheet
    path = path_to_excel_template
    wb = load_workbook(path, read_only=True)
    warnings.simplefilter(action='ignore', category=UserWarning)

    sheet = wb["OpenProjectData"]

    # Add data to a dictionary of dictionaries for easier manipulation
    sprintdict = {}

    for row in sheet.iter_rows(min_row=6, min_col=15, max_row=sheet.max_row, max_col=15):
        for cell in row:
            if cell.value is not None:
                sprintdict.update({cell.value: {
                    "Story": 0,
                    "AEM Task": 0,
                    "Bug": 0,
                    "Feature": 0,
                }})

    current_row = 5

    # Iterate through Excel file and check for every sprint their issues type and then count them
    for row in sheet.iter_rows(min_row=6, min_col=4, max_row=sheet.max_row, max_col=15):
        current_row = current_row + 1
        for cell in row:
            for sprintName in sprintdict:
                if cell.value == sprintName and cell.value is not None:
                    sprint = sprintdict[sprintName]
                    if sheet.cell(current_row, 4).value == 'Story':
                        sprint["Story"] = sprint["Story"] + 1
                    elif sheet.cell(current_row, 4).value == 'AEM Task':
                        sprint["AEM Task"] = sprint["AEM Task"] + 1
                    elif sheet.cell(current_row, 4).value == 'Bug':
                        sprint["Bug"] = sprint["Bug"] + 1
                    elif sheet.cell(current_row, 4).value == 'Feature':
                        sprint["Feature"] = sprint["Feature"] + 1

    # Save data from dictionary in a list to use in the chart
    sprint_names = []
    state_dict = {}
    for sprintName in sprintdict:
        sprint = sprintdict[sprintName]
        sprint_names.append(sprintName)
        for state in sprint:
            try:
                state_dict[state].append(sprint[state])
            except:
                state_dict[state] = [sprint[state]]

    # Use data saved from dictionaries to plot bar chart
    plotdata = pd.DataFrame(state_dict, sprint_names)
    X_axis = np.arange(len(sprint_names))
    plotdata.plot(kind="bar", stacked=True, figsize=(20, 15))
    labels = ['\n'.join(wrap(l, 20)) for l in sprint_names]
    plt.rcParams.update({'font.size': 15})
    plt.xticks(X_axis, labels, rotation=90)
    plt.title("Count state for each sprint")
    plt.xlabel("Sprint")
    plt.ylabel("State count")
    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
    plt.tight_layout()

    plt.savefig(path_to_output_folder + '\sprint_type.png')
    plt.close()


def work_priority_date(path_to_excel_template, path_to_csv_file, path_to_output_folder):

    # Load template Excel file and pick the OpenProjectData sheet
    path = path_to_excel_template
    wb = load_workbook(path, read_only=True)
    warnings.simplefilter(action='ignore', category=UserWarning)

    sheet = wb["OpenProjectData"]

    current_date = datetime.today().strftime('%Y-%m-%d')

    # Add data to a dictionary of dictionaries for easier manipulation
    datedict = {}

    datedict.update({current_date: {
        "Low": 0,
        "Normal": 0,
        "High": 0
    }})

    # Iterate through Excel file and check for current date how many open tasks remain and their priority
    current_row = 5
    current_col = 3
    for dateName in datedict:
        if dateName == current_date:
            date = datedict[dateName]
            for row in sheet.iter_rows(min_row=6, min_col=7, max_row=sheet.max_row, max_col=7):
                current_row += 1
                current_col += 1
                for cell in row:
                    if cell.value == "Normal" and sheet.cell(current_row, 6).value != "Closed":
                        date["Normal"] = date["Normal"] + 1
                    elif cell.value == "Low" and sheet.cell(current_row, 6).value != "Closed":
                        date["Low"] = date["Low"] + 1
                    elif cell.value == "High" and sheet.cell(current_row, 6).value != "Closed":
                        date["High"] = date["High"] + 1

    # Save data from dictionary in a list to use in the chart
    date_names = []
    date_dict = {}
    for dateName in datedict:
        date = datedict[dateName]
        date_names.append(dateName)
        for state in date:
            try:
                date_dict[state].append(date[state])
            except:
                date_dict[state] = [date[state]]

    # Prepare data to be stored in a csv file
    headers = ["Low", "Normal", "High"]
    plotdata = pd.DataFrame(date_dict, date_names)

    # Append data to the csv file, if the file does not exist, create it first
    if os.path.isfile(os.path.normpath(os.path.join(path_to_csv_file, "priority_for_each_day.csv"))):
        plotdata.to_csv(os.path.join(path_to_csv_file, "priority_for_each_day.csv"), mode='a', header=False)
    else:
        plotdata.to_csv(os.path.join(path_to_csv_file, "priority_for_each_day.csv"), mode='w', header=headers)

    # Read CSV only for getting it's data length
    with open(os.path.join(path_to_csv_file, 'priority_for_each_day.csv'), newline='') as f:
        reader = csv.reader(f)
        csv_list = list(reader)

    # Keep only the last 60 entries on the chart
    data_to_be_read = pd.read_csv(os.path.normpath(os.path.join(path_to_csv_file, "priority_for_each_day.csv")), skiprows=[skip_row for skip_row in range(1,len(csv_list)-60)])

    # Plot chart using data from the csv file
    df = pd.DataFrame(data_to_be_read)
    X = list(df.iloc[:, 0])
    X_axis = np.arange(len(X))
    df.plot(kind="bar", stacked=True, figsize=(20, 15))
    plt.rcParams.update({'font.size': 15})
    plt.xticks(X_axis, X, rotation=90)
    plt.title("Count open issues priority for each date")
    plt.xlabel("Date")
    plt.ylabel("Priority count")
    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
    plt.tight_layout()

    plt.savefig(os.path.normpath(os.path.join(path_to_output_folder, 'date_priority.png')))
    plt.close()


def main():

    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input-path-excel-template', help="Path to .xlsm excel template file. Path must also contain the filename and extension.", required=True)
    parser.add_argument('-t', '--input-path-excel-library', help="Path to .xlam excel library file. Path must also contain the filename and extension", required=True)
    parser.add_argument('-l', '--output-path-csv', help="Path to the csv file where data about issues's priority each day is stored.", required=True)
    parser.add_argument('-o', '--output-folder', help="Path to output folder with all generated graphs.", required=True)
    args = parser.parse_args()

    path_to_excel_template = args.input_path_excel_template
    path_to_excel_library = args.input_path_excel_library
    path_to_csv_file = args.output_path_csv
    path_to_output_folder = args.output_folder

    # task kill excel
    os.system('Taskkill /IM EXCEL.EXE /F')

    print("Running script")
    print("==========================================")
    # Start Excel application
    excel1 = win32.Dispatch("Excel.Application")
    excel2 = win32.Dispatch("Excel.Application")

    # Get the path of the Excel library and Excel template file
    path1 = path_to_excel_library
    path2 = path_to_excel_template

    # Open the Excel library and Excel template
    wb1 = excel1.Workbooks.Open(path1)
    wb2 = excel2.Workbooks.Open(path2)

    # Hide both Excel files
    excel1.Visible = False
    excel2.Visible = False

    # Activate only the worksheet where we will collect our data
    excel2.Worksheets("OpenProjectData").Activate()

    # Use the button to clear old data and collect the one from OpenProject
    try:
        print("Fetching Data...")
        excel2.Application.Run("API_ClearAndLoadAll")
    except:
        print("Failed to load data")
        pass

    # Save the data to the Excel template
    excel2.ActiveWorkbook.Save()
    print("==========================================")
    print("Data fetched")
    print("==========================================")
    print("Processing data...")
    print("==========================================")

    # Close both Excel library and Excel template

    # task kill excel
    # os.system('Taskkill /IM EXCEL.EXE /F')

    try:
        excel2.Quit()
    except:
        pass
    try:
        excel1.Quit()
    except:
        pass

    # Execute all functions written in this script
    story_task_data(path_to_excel_template, path_to_output_folder)
    bug_feature_data(path_to_excel_template, path_to_output_folder)
    user_data(path_to_excel_template, path_to_output_folder)
    sprint_status_data(path_to_excel_template, path_to_output_folder)
    sprint_type_data(path_to_excel_template, path_to_output_folder)
    work_priority_date(path_to_excel_template, path_to_csv_file, path_to_output_folder)
    print("Done processing data")
    print("==========================================")
    print("Script ran successfully")
    print("==========================================")

if __name__ == "__main__":
    main()
