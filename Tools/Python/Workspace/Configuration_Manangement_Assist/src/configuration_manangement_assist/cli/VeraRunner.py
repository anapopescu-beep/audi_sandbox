"""VeraRunner.py: Used to create statistics for Jenkins Integration. Check YAML Configuration file."""

__author__ = 'Mihai Motoc'
__copyright__ = "Copyright 2023, Autoliv Electronic"
__version__ = "1.0.0"
__email__ = 'mihai.motoc@autoliv.com'
__status__ = 'Released'

import yaml
import os
import subprocess
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import *
import matplotlib.pyplot as plt
import argparse


def main():
    global yaml_config

    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input-yaml-config-path', required=True)
    args = parser.parse_args()
    yaml_config_path = args.input_yaml_config_path

    list_of_source_files = []
    dict_of_files = {}

    # yaml_config_path = "VeraRunnerConfigFile.yaml"
    yaml_config = yaml.safe_load(open(yaml_config_path, 'r').read())
    project_path = yaml_config['project_path']
    vera_path = yaml_config['vera_path']
    output_path = yaml_config['output_path']
    list_of_directories_to_ignore = yaml_config['list_of_directories_to_ignore']

    for root, dirnames, filenames in os.walk(project_path):
        for filename in filenames:
                if "Implementation" in root:
                    if filename.endswith('.c') or filename.endswith('.h'):
                        sw_component = root.split('\\')[-3]
                        if sw_component not in list_of_directories_to_ignore:
                            list_of_source_files.append(filename)
                            if sw_component not in dict_of_files.keys():
                                dict_of_files[sw_component] = {}
                            dict_of_files[sw_component][filename] = {'file_path': root + "\\" + filename}

    for swComp in dict_of_files.keys():
        for file in dict_of_files[swComp].keys():
            path = dict_of_files[swComp][file]['file_path']
            console_output = subprocess.check_output([vera_path, path]).decode()
            list_console_output = console_output.split('\n')[2:]
            dict_of_files[swComp][file] = {'console_output': list_console_output}

    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    sheet.title = 'Vera Warnings'

    # Set Start Row Where you want to write all values
    start_row = 4
    # Set a big and nice title at the beggining
    # ==========================================================
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=6)
    sheet.cell(row=1, column=1).value = 'VERA WARNINGS'
    sheet.cell(row=1, column=1).font = Font(bold=True, size=18)
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=1).font = Font(bold=True)
    sheet.cell(row=start_row, column=2).alignment = Alignment(horizontal='center')
    sheet.cell(row=start_row, column=2).font = Font(bold=True)

    sheet.freeze_panes = 'A' + str(start_row + 1)
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 100

    row_count = start_row + 1
    warning_counter = 0

    for module in dict_of_files.keys():
        sheet.merge_cells(start_row=row_count, start_column=1, end_row=row_count, end_column=8)
        sheet.cell(row=row_count, column=1).alignment = Alignment(horizontal='center')
        sheet.cell(row=row_count, column=1).font = Font(bold=True)
        sheet.cell(row=row_count, column=1).value = module
        row_count += 1

        for file in dict_of_files[module].keys():
            for warning in dict_of_files[module][file]['console_output']:
                try:
                    sheet.cell(row=row_count, column=1).value = file
                    sheet.cell(row=row_count, column=2).value = "line" + warning.split(file)[1]
                    row_count += 1
                    warning_counter += 1
                except:
                    print("Error " + file + " " + warning)

    sheet_graphs = wb.create_sheet("Graphics")
    sheet_graphs.column_dimensions['A'].width = 20
    sheet_graphs.column_dimensions['B'].width = 20
    sheet_graphs.column_dimensions['C'].width = 20
    sheet_graphs.column_dimensions['D'].width = 20
    sheet_graphs.column_dimensions['E'].width = 20

    data = {}
    warning_data = {}
    modules = []
    for module in dict_of_files:
        modules.append(module)
        for file, file_data in dict_of_files[module].items():
            warnings_count = len(file_data['console_output'])
            if module not in data:
                data[module] = {}
            data[module][file] = warnings_count
            warning_data[module] = warnings_count

    plot_dict = {}
    for module in dict_of_files.keys():
        plot_dict[module] = 0  # Initialize the sum to zero
        for file, file_data in dict_of_files[module].items():
            plot_dict[module] += len(file_data['console_output'])

    row_num = 1
    for module, files in dict_of_files.items():
        # Write the module name in the first cell of the row
        sheet_graphs.cell(row=row_num, column=3, value=module).fill = PatternFill(start_color='c2fc03', end_color='c2fc03', fill_type="solid")
        row_num += 1

        # Write the files' names in a row
        files_row = row_num
        for col_num, file in enumerate(files, start=1):
            sheet_graphs.cell(row=row_num, column=col_num, value=file).fill = PatternFill(start_color='fca903', end_color='fca903', fill_type="solid")

        row_num += 1

        # Write the values in the row below
        values_row = row_num
        for col_num, file_data in enumerate(files.values(), start=1):
            warnings_count = len(file_data['console_output'])
            sheet_graphs.cell(row=row_num, column=col_num, value=warnings_count).fill = PatternFill(start_color='fc4103', end_color='fc4103', fill_type="solid")

        row_num += 1

        # Add an empty row after each module's data
        row_num += 1

    modules = list(plot_dict.keys())
    counts = list(plot_dict.values())
    bar_width = 0.5
    fig, ax = plt.subplots(figsize=(30, 6))

    plt.bar(modules, counts, width=bar_width)

    plt.xlabel('Modules')
    plt.ylabel('Number of warnings')
    plt.title('Number of warnings for Each Module')

    chart_image_path = os.path.join(output_path, 'vera_warnings.png')

    plt.savefig(chart_image_path)
    plt.close()

    # Get the active sheet
    sheet = sheet_graphs

    # Load the chart image into Excel
    img = openpyxl.drawing.image.Image(os.path.join(output_path, chart_image_path))
    img.anchor = 'I1'  # Set the anchor cell where the chart should be placed
    sheet.add_image(img)
    plt.show()

    wb.save(os.path.join(output_path, 'VeraWarnings.xlsx'))


if __name__ == '__main__':
    main()


