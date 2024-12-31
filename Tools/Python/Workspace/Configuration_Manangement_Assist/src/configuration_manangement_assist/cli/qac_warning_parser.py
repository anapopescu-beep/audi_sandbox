"""qac_warning_parser.py: Gathers information regarding QAC reports for Eclipse"""

__author__ = 'Roxana Dimanescu'
__copyright__ = "Copyright 2023, Autoliv Electronic"
__version__ = "1.0.0"
__email__ = 'roxana.dimanescu@autoliv.com'
__status__ = 'Released'

import os
import openpyxl
import argparse
from pathlib import Path

def main():

    #Creat the parser
    parser = argparse.ArgumentParser()

    #Adding the arguments to parser
    parser.add_argument('-i', '--path_folder', required=True)
    args = parser.parse_args()
    path_folder = args.path_folder
    initial_path_folder = Path(path_folder)

    #Set in the variable the folder name
    name_file = initial_path_folder.name

    #Set the path for the folder parent folder
    component_path_folder = initial_path_folder.parents[2]

    #Initializat the variable for the precondition
    found_folder = False
    found_file = False

    #Initializat the variable for the  path with "None"
    path_folder_excel = None
    path_file = None

    #Search from the folder in the parent path to find the folder where excel file for the ".c" file are
    for root, dirs, files in os.walk(component_path_folder):
        for dir in dirs:
            if "Static_Analysis" in dir:
                #save the path for the excel file in the variable
                path_folder_excel = os.path.join(root, dir)
                #print(path_folder_excel)
                found_folder = True
                break

    #Search from the excel file for the ".c" file in the folder
    if found_folder == True:
        for root, dirs, files in os.walk(path_folder_excel):
            for file in files:
                if (file.endswith(".xlsx") and '_WarningReport' in file and name_file in file):
                    path_file = os.path.join(root, file)
                    found_file = True
                    break

    #Don't exist the folder in the path
    #elif found_folder == False:
        #print("Warning : Don't find the folder!")

    #Don't exist the file in the folder "Static_Analysis"
    #if path_file == None:
        #print("Warning : Don't find the file!")

    #Print in the console the rows for the excel with the column : linie, type, severity, warning code, worning message, source
    if found_file == True:
        wb_obj = openpyxl.load_workbook(path_file)
        sheet_obj = wb_obj.active
        rows = sheet_obj.max_row
        #rows += 1

        if rows > 5:
            #print("Warning: The file don't have values")
        #else:
            my_list = list()
            for value in sheet_obj.iter_rows(min_row=5, max_row=rows, min_col=1, max_col=5, values_only=True):
                row = []
                for i, v in enumerate(value):
                    if i == 3:  # Ignorarea coloanei 4
                        continue
                    row.append(v)
                my_list.append(row)
        for item in my_list:
            #print(name_file, *item )
            output =  ';' + str(item[0]) + '; '
            output += ' '.join(str(x) for x in item[1:])

            print(path_folder + output)

if __name__ == '__main__':
    main()